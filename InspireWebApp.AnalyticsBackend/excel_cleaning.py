import math
import os.path
from typing import Dict

import pandas as pd
import openpyxl
import unidecode
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame

file_path = 'Sales Report List.xlsx'

stage_1_file_path = 'Stage 1 Sales Report List.xlsx'

stage_2_file_path = 'Stage 2 Sales Report List.csv'

LETTER_TO_INDEX_MAPPINGS: Dict[chr, int] = {chr(i): i - ord('A') for i in range(ord('A'), ord('Z') + 1)}


def unmerge_cells(sheet: Worksheet):
    merged_cells = list(sheet.merged_cells.ranges)

    if len(merged_cells) == 0:
        return

    # Unmerge all merged cells
    for merged_cell in merged_cells:
        sheet.unmerge_cells(str(merged_cell))


def shift_data(sheet: Worksheet, from_col: chr, to_col: chr):

    from_idx = LETTER_TO_INDEX_MAPPINGS[from_col]
    to_idx = LETTER_TO_INDEX_MAPPINGS[to_col]

    # Iterate through each row in the sheet
    for row in sheet.iter_rows(min_row=15):
        # Get the value from column O
        value_from_O = row[from_idx].value
        # Clear the value from column O
        row[from_idx].value = None
        # Paste the value to column P
        row[to_idx].value = value_from_O


# Function to autofit column width according to the longest value in each cell for all columns
def autofit_columns(sheet: Worksheet):
    for col in sheet.columns:
        column = col[0].column_letter  # Get the column letter

        all_values = (len(str(cell.value)) for cell in col if cell.value)
        max_length = max(all_values, default=0)

        # Set the column width based on the longest cell value
        sheet.column_dimensions[column].width = max_length + 2  # Add a little padding


def set_headers(df: DataFrame) -> DataFrame:
    # The first row contains the headers
    column_names = df.iloc[0]

    # Rename the columns using a dictionary comprehension
    # Ignore the columns that have nan as their name
    mappings = {old_name: new_name for old_name, new_name in zip(df.columns, column_names) if isinstance(new_name, str)}

    df = df.rename(columns=mappings)

    # Our data records start from the fourth record so drop the first three rows of the DataFrame
    df.drop(df.index[:3], axis=0, inplace=True)
    return df


def prepare_for_pandas(src_path: str, dest_path: str):
    # Load the Excel file
    workbook = openpyxl.load_workbook(src_path)

    # The sheet is not in an acceptable format to be inserted into the database
    sheet = workbook[workbook.sheetnames[0]]

    # Un-squash all merged cells
    unmerge_cells(sheet)

    # For an odd reason, after the un-merge the data of column P has shifted to column O,
    # so we shift the data back to P column
    shift_data(sheet, 'O', 'P')

    # Auto-fit columns for better readability during debugging
    autofit_columns(sheet)

    # Save the modified workbook to a new file (or overwrite the original file)
    workbook.save(dest_path)


def sanitize_description(description: str):

    if description.find('"') != -1:
        description = description.replace('"', '')

    if description.find(',') != -1:
        description = description.replace(',', '.')

    if description.find('  ') != -1:
        description = description.replace('  ', ' ')

    from unidecode import unidecode

    return unidecode(description)


def prepare_for_db_import(src_path: str, dest_path: str):
    # The headers are on the 10th row, so ignore earlier rows
    df = pd.read_excel(src_path, header=10)

    df = set_headers(df)

    #
    df.drop(df.columns[1:3], axis=1, inplace=True)

    # The second column corresponds to the customers
    df = df.rename(columns={df.columns[1]: 'Customer'})

    # Drop all empty rows
    df = df.dropna(how='all')

    # Drop all columns that do not have any values
    df = df.dropna(how='all', axis=1)

    # Fill in missing customer names
    current_customer_name = df.iloc[0]['Customer']
    for index, row in df.iterrows():
        customer_name = row['Customer']
        if not isinstance(customer_name, str) and math.isnan(customer_name):
            row['Customer'] = current_customer_name
        else:
            current_customer_name = customer_name

    df = df[df.iloc[:, 0].notnull()]  # Keep only rows where the first column is not null

    # Replace all commas with dots
    df['Description'] = df['Description'].apply(sanitize_description)

    # Write the DataFrame to an Excel file
    df.to_csv(dest_path, index=False)


def process_excel_file():

    if not os.path.exists(stage_1_file_path):
        prepare_for_pandas(file_path, stage_1_file_path)

    prepare_for_db_import(stage_1_file_path, stage_2_file_path)


if __name__ == '__main__':
    process_excel_file()

